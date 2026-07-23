"""Curated-list import (XLSX/CSV) for Manual Scan — a bounded, security-hardened file-parse layer.

Turns an uploaded .xlsx/.csv into canonical seed domains + display-only metadata and a disposition
against the EXISTING ``AnalyzedSiteRegistry``. It never persists the workbook, never evaluates a
formula or macro, treats every cell as untrusted text, and enforces hard ceilings (size / rows /
cols / sheets). Bare domains become ``https://<domain>`` seeds via the shared ``canonical_domain``.
Launch reuses the existing manual Scout path — this module produces seeds; it never runs a campaign.
"""
from __future__ import annotations

import csv
import hashlib
import io
import re
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from core.scout.discovery.analyzed_registry import ANALYZED, REJECTED, AnalyzedSiteRegistry
from core.scout.discovery.domain_intel import canonical_domain

# Header names (case-insensitive) that identify the seed column, in priority order.
_URL_HEADERS = ("scout seed url", "seed url", "url", "website", "domain", "seed")
_MAX_META_COLS = 12
_MAX_CELL = 200
# A STRICT domain shape on top of canonical_domain (which is lenient): labels of a-z0-9-, at least one
# dot, an alphabetic TLD >= 2. Untrusted cells (formulas, prose, junk) that canonical_domain may half-
# parse are rejected here so a curated list can never seed a bogus target.
_DOMAIN_RE = re.compile(r"^(?=.{1,253}$)([a-z0-9]([a-z0-9-]{0,61}[a-z0-9])?\.)+[a-z]{2,}$")


def _plausible_domain(dom: str) -> bool:
    return bool(_DOMAIN_RE.match(dom))


class CuratedImportError(ValueError):
    """Raised when an uploaded curated list is unsupported, malformed, oversized, or has no URL column."""


@dataclass
class ImportRow:
    row: int
    original: str
    canonical_domain: str
    seed_url: str
    valid: bool
    disposition: str                 # new | already_analyzed | rejected | duplicate | invalid
    recommended_action: str = ""     # from a "Recommended action" column (e.g. "Scout now" / "Backlog")
    preselect: bool = False          # whether the UI should pre-check this row (see _preselect)
    metadata: Dict[str, str] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, Any]:
        return dict(self.__dict__)


@dataclass
class ImportResult:
    import_id: str
    kind: str                        # xlsx | csv
    column: str                      # detected seed-column header
    rows: List[ImportRow]
    counters: Dict[str, int]
    has_recommended_action: bool = False   # True when the file had a "Recommended action" column

    def to_dict(self) -> Dict[str, Any]:
        return {"import_id": self.import_id, "kind": self.kind, "column": self.column,
                "rows": [r.to_dict() for r in self.rows], "counters": self.counters,
                "has_recommended_action": self.has_recommended_action}


def _now_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def _grid_from_csv(data: bytes, max_rows: int, max_cols: int) -> List[List[str]]:
    text = data.decode("utf-8", errors="replace")
    grid: List[List[str]] = []
    try:
        # row 0 is the header; rows 1..max_rows are the allowed data rows (max_rows data rows).
        for i, row in enumerate(csv.reader(io.StringIO(text))):
            if i > max_rows:
                raise CuratedImportError(f"too many rows (limit {max_rows})")
            if len(row) > max_cols:
                raise CuratedImportError(f"too many columns (limit {max_cols})")
            grid.append([("" if c is None else str(c)) for c in row])
    except CuratedImportError:
        raise
    except Exception as exc:  # noqa: BLE001 - a malformed CSV (e.g. NUL bytes) is a refusal, not a crash
        raise CuratedImportError(f"unreadable CSV: {type(exc).__name__}") from exc
    return grid


def _guard_zip_bomb(data: bytes, *, max_uncompressed: int = 64 * 1024 * 1024,
                    max_ratio: int = 200) -> None:
    """An .xlsx IS a zip. Read the central-directory member sizes (no decompression) and refuse a
    decompression bomb BEFORE openpyxl parses it — the upload-size limit alone does not bound the
    decompressed size."""
    import zipfile
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            infos = zf.infolist()
    except (zipfile.BadZipFile, OSError) as exc:
        raise CuratedImportError("not a valid .xlsx (zip) container") from exc
    total = sum(int(getattr(i, "file_size", 0)) for i in infos)
    comp = sum(int(getattr(i, "compress_size", 0)) for i in infos) or 1
    if total > max_uncompressed:
        raise CuratedImportError(f"workbook decompresses too large (> {max_uncompressed} bytes)")
    if total // comp > max_ratio:
        raise CuratedImportError("workbook compression ratio too high (possible zip bomb)")


def _grid_from_xlsx(data: bytes, max_rows: int, max_cols: int, max_sheets: int) -> List[List[str]]:
    _guard_zip_bomb(data)                                 # bound decompression before openpyxl runs
    try:
        import openpyxl
        # read_only + data_only: a formula is NEVER evaluated; the presented cached value / text is
        # returned and later validated as data (never trusted just because a cell exists).
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except CuratedImportError:
        raise
    except Exception as exc:  # noqa: BLE001 - any parse/decrypt failure is a refusal, never a crash
        raise CuratedImportError(f"unreadable or encrypted workbook: {type(exc).__name__}") from exc
    if len(wb.sheetnames) > max_sheets:
        raise CuratedImportError(f"too many sheets (limit {max_sheets})")
    ws = wb[wb.sheetnames[0]]
    grid: List[List[str]] = []
    try:
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > max_rows:
                raise CuratedImportError(f"too many rows (limit {max_rows})")
            if len(row) > max_cols:
                raise CuratedImportError(f"too many columns (limit {max_cols})")
            grid.append([("" if c is None else str(c)) for c in row])
    except CuratedImportError:
        raise
    except Exception as exc:  # noqa: BLE001 - a corrupt lazy read is a refusal, not a crash
        raise CuratedImportError(f"unreadable workbook rows: {type(exc).__name__}") from exc
    return grid


def _detect_column(header: List[str]) -> int:
    norm = [h.strip().lower() for h in header]
    for name in _URL_HEADERS:                             # priority order
        if name in norm:
            return norm.index(name)
    raise CuratedImportError("no URL/Website/Domain/'Scout seed URL' column found")


def parse_curated_list(data: bytes, filename: str, *, registry: Optional[AnalyzedSiteRegistry] = None,
                       max_bytes: int = 5_000_000, max_rows: int = 500, max_cols: int = 40,
                       max_sheets: int = 5) -> ImportResult:
    """Parse an uploaded curated list into canonical seed rows. Raises CuratedImportError on any
    unsupported/malformed/oversized/ambiguous input. Never persists the file or evaluates a cell."""
    if not isinstance(data, (bytes, bytearray)) or not data:
        raise CuratedImportError("empty upload")
    if len(data) > max_bytes:
        raise CuratedImportError(f"file too large (limit {max_bytes} bytes)")
    ext = ("." + filename.rsplit(".", 1)[-1].lower()) if "." in filename else ""
    if ext == ".csv":
        kind, grid = "csv", _grid_from_csv(bytes(data), max_rows, max_cols)
    elif ext == ".xlsx":
        kind, grid = "xlsx", _grid_from_xlsx(bytes(data), max_rows, max_cols, max_sheets)
    else:
        raise CuratedImportError(f"unsupported file type {ext or '(none)'} — use .xlsx or .csv")
    if not grid:
        raise CuratedImportError("no rows in file")
    header = grid[0]                                       # per-row col/row bounds enforced in readers
    col = _detect_column(header)
    norm_hdr = [h.strip().lower() for h in header]
    rec_idx = norm_hdr.index("recommended action") if "recommended action" in norm_hdr else -1
    has_rec = rec_idx >= 0

    rows: List[ImportRow] = []
    seen: set = set()
    counters = {"total": 0, "valid_unique": 0, "invalid": 0, "dup_in_file": 0,
                "already_analyzed": 0, "rejected": 0, "preselected": 0}
    for idx, raw in enumerate(grid[1:], start=2):
        counters["total"] += 1
        original = (raw[col] if col < len(raw) else "").strip()[:_MAX_CELL]
        dom = canonical_domain(original).lower()
        rec = (str(raw[rec_idx]).strip()[:_MAX_CELL] if 0 <= rec_idx < len(raw) else "")
        meta = {header[j].strip()[:_MAX_CELL]: str(raw[j]).strip()[:_MAX_CELL]
                for j in range(min(len(header), _MAX_META_COLS)) if j != col and j < len(raw)
                and header[j].strip()}
        if not _plausible_domain(dom):                    # not a real domain -> invalid, never run
            counters["invalid"] += 1
            rows.append(ImportRow(idx, original, "", "", False, "invalid",
                                  recommended_action=rec, preselect=False, metadata=meta))
            continue
        if dom in seen:
            counters["dup_in_file"] += 1
            rows.append(ImportRow(idx, original, dom, f"https://{dom}", False, "duplicate",
                                  recommended_action=rec, preselect=False, metadata=meta))
            continue
        seen.add(dom)
        disposition = "new"
        if registry is not None:
            entry = registry.get(dom)
            status = getattr(entry, "analysis_status", "") if entry else ""
            if status == REJECTED:
                disposition = "rejected"
                counters["rejected"] += 1
            elif status == ANALYZED:
                disposition = "already_analyzed"
                counters["already_analyzed"] += 1
        if disposition == "new":
            counters["valid_unique"] += 1
        pre = _preselect(disposition, has_rec, rec)
        if pre:
            counters["preselected"] += 1
        rows.append(ImportRow(idx, original, dom, f"https://{dom}", True, disposition,
                              recommended_action=rec, preselect=pre, metadata=meta))

    import_id = f"imp-{_now_stamp()}-{hashlib.sha1(bytes(data)).hexdigest()[:8]}"
    return ImportResult(import_id=import_id, kind=kind, column=header[col], rows=rows,
                        counters=counters, has_recommended_action=has_rec)


def _preselect(disposition: str, has_rec: bool, rec: str) -> bool:
    """Pre-check ONLY a scannable NEW target. When a 'Recommended action' column exists, honour it —
    only 'Scout now' is preselected; Backlog / Exclude stay unchecked. Otherwise every valid-new row is
    preselected. already_analyzed / rejected / invalid / duplicate are never preselected."""
    if disposition != "new":
        return False
    if has_rec:
        return rec.strip().lower() == "scout now"
    return True
